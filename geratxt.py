#!/usr/bin/env python3
"""
geratxt-clean.py

- Lê o Excel com openpyxl (modo normal, sem converter '-')
- Extrai apenas colunas Bola1..Bola20
- Substitui '-' ou valores não numéricos por vazio
- Grava em TXT sem cabeçalho, separado por TAB
"""

import sys
import re
import csv
from pathlib import Path
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print("Uso: python geratxt-clean.py Lotomania.xlsx")
        sys.exit(1)

    in_file = Path(sys.argv[1])
    out_file = in_file.with_name(in_file.stem + "_Bolas.txt")

    # abre workbook em modo normal
    wb = load_workbook(filename=in_file, read_only=False, data_only=True)
    ws = wb["LOTOMANIA"]

    # pega cabeçalho
    header_row = next(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in header_row]

    # detecta colunas Bola1..Bola20
    pattern = re.compile(r"^Bola\d+$", re.IGNORECASE)
    indices = [i for i, h in enumerate(header) if pattern.match(h)]

    if not indices:
        print("[erro] Não achei colunas BolaN")
        sys.exit(2)

    # escreve TXT
    with open(out_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        for row in ws.iter_rows(values_only=True, min_row=2):  # pula cabeçalho
            values = []
            for i in indices:
                val = row[i] if i < len(row) else ""
                if val in (None, "-"):  # trata None e '-' como vazio
                    values.append("")
                else:
                    values.append(str(val))
            if any(values):  # ignora linhas totalmente vazias
                writer.writerow(values)

    print(f"[ok] Arquivo gerado: {out_file}")

if __name__ == "__main__":
    main()
