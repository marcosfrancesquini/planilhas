#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ordenar_planilhas_formula_safe.py

- Multi-input (-i repetido ou separado por vírgula).
- Lê valores CALCULADOS de fórmulas (XLSX/XLSM via openpyxl data_only=True).
- --recalc-excel: abre no Excel (xlwings), recalcula e salva antes de ler.
- --debug: gera CSV por aba mostrando o que veio nas colunas-alvo.

Saída: 1 aba única contendo, para cada aba de entrada:
  "arquivo :: Nome da Planilha | Tamanho do bloco: x | Tamanho do ultimo bloco: y | Porcentagem Faltante: z% | Número de linhas faltantes: k"
  + 7 blocos lado a lado; cada bloco tem 3 colunas: <critério> | numero base | Número.
"""
import argparse, re, unicodedata, os
from pathlib import Path
from typing import List, Tuple, Any, Dict, Optional
import numpy as np
import pandas as pd

CRITERIOS: List[Tuple[str, bool]] = [
    ("Máxima Freq. nos Blocos Completos", False),
    ("Mínima Freq. nos Blocos Completos", True),
    ("Frequência no Último Bloco Incompleto", False),
    ("Maximo de vezes", False),
    ("Minimo de vezes", True),
    ("máximo grande", False),
    ("mínimo pequeno", True),
]

SYNONYMS: Dict[str, List[str]] = {
    "Máxima Freq. nos Blocos Completos": ["Maxima Freq. nos Blocos Completos","Maxima Freq nos Blocos Completos","Máxima Freq nos Blocos Completos"],
    "Mínima Freq. nos Blocos Completos": ["Minima Freq. nos Blocos Completos","Minima Freq nos Blocos Completos","Mínima Freq nos Blocos Completos"],
    "Frequência no Último Bloco Incompleto": ["Frequencia no Ultimo Bloco Incompleto","Frequencia no Último Bloco Incompleto"],
    "Maximo de vezes": ["Máximo de vezes","Maximo de Vezes"],
    "Minimo de vezes": ["Mínimo de vezes","Minimo de Vezes"],
    "máximo grande": ["Maximo grande","maximo grande"],
    "mínimo pequeno": ["Minimo pequeno","minimo pequeno"],
}

# ---------- util ----------
def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def _build_col_resolver(df: pd.DataFrame) -> Dict[str, str]:
    return {_norm_text(c): c for c in df.columns}

def _resolve_col(df: pd.DataFrame, canonical: str, resolver: Dict[str, str], extras: Optional[List[str]] = None) -> Optional[str]:
    candidates = [canonical] + (extras or [])
    for cand in candidates:
        key = _norm_text(cand)
        if key in resolver:
            return resolver[key]
    key = _norm_text(canonical).replace("freq","frequencia")
    return resolver.get(key)

def _first_non_null(df: pd.DataFrame, candidates: List[str]) -> Any:
    for c in candidates:
        if c in df.columns:
            s = pd.Series(df[c]).dropna()
            if len(s) > 0:
                return s.iloc[0]
    return np.nan

def _fmt_percent(val: Any) -> str:
    if pd.isna(val): return ""
    try:
        s = str(val).strip().replace("%","").replace(",",".")
        return f"{float(s):.2f}%"
    except Exception:
        return str(val)

def _pick_numero_base(df: pd.DataFrame) -> pd.Series:
    if "numero base" in df.columns: return df["numero base"]
    for alt in ["Número","numero","numero_base","número base"]:
        if alt in df.columns:
            s = df[alt].copy(); s.name = "numero base"; return s
    return pd.Series(df.index + 1, index=df.index, name="numero base")

def _pick_numero_col(df: pd.DataFrame) -> pd.Series:
    for alt in ["Número","numero","num","Numero","NÚMERO","número"]:
        if alt in df.columns:
            s = df[alt].copy(); s.name = "Número"; return s
    nb = _pick_numero_base(df); s = nb.copy(); s.name = "Número"; return s

def _coerce_numeric_col(series: pd.Series) -> pd.Series:
    if series.dtype != object: return series
    try:
        return pd.to_numeric(series.astype(str).str.replace("%","",regex=False).str.replace(",",".",regex=False).str.strip(), errors="coerce")
    except Exception:
        return series

# ---------- leitura com fórmulas ----------
def read_sheet_values(path: str, sheet_name: str) -> pd.DataFrame:
    p = Path(path); ext = p.suffix.lower()
    if ext in [".xlsx",".xlsm"]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames: return pd.DataFrame()
            ws = wb[sheet_name]
            rows = list(ws.values)
            if not rows: return pd.DataFrame()
            header = [f"Unnamed: {i}" if (h in (None,"")) else h for i,h in enumerate(rows[0])]
            data = rows[1:]
            return pd.DataFrame(data, columns=header)
        except Exception:
            return pd.read_excel(path, sheet_name=sheet_name)
    else:
        # ODS / XLS antigo: pandas normal (não calcula fórmulas)
        return pd.read_excel(path, sheet_name=sheet_name)

# ---------- blocos / escrita ----------
def _build_blocks_side_by_side(df: pd.DataFrame, zero_as_blank: bool) -> Dict[str, pd.DataFrame]:
    nb = _pick_numero_base(df)
    numcol = _pick_numero_col(df)
    blocks: Dict[str, pd.DataFrame] = {}
    resolver = _build_col_resolver(df)

    for crit, asc in CRITERIOS:
        real_col = _resolve_col(df, crit, resolver, SYNONYMS.get(crit))
        if real_col is None:
            ord_df = pd.DataFrame({crit: pd.Series(dtype=object), "numero base": pd.Series(dtype=object), "Número": pd.Series(dtype=object)})
        else:
            orig = df[real_col].copy()
            if zero_as_blank and crit in ["máximo grande","mínimo pequeno"]:
                if pd.api.types.is_numeric_dtype(orig):
                    orig = orig.mask(orig == 0, other=np.nan)
                else:
                    orig = orig.mask(orig.astype(str).str.strip().isin({"0","0.0","0,0"}), other=np.nan)
            key = _coerce_numeric_col(df[real_col])
            ord_df = pd.DataFrame({crit: orig, "numero base": nb, "Número": numcol, "__key__": key})
            ord_df = ord_df.sort_values(by="__key__", ascending=asc, kind="mergesort", na_position="last").drop(columns=["__key__"])
        blocks[crit] = ord_df.reset_index(drop=True)
    return blocks

def _xlsx_value(v):
    try:
        if v is None: return None
        if isinstance(v,(float,int)) and (pd.isna(v) or (isinstance(v,float) and (np.isnan(v) or np.isinf(v)))): return None
        if pd.isna(v): return None
        return v
    except Exception:
        return v

def _write_section(worksheet, start_row: int, header_cells: List[str], blocks: Dict[str, pd.DataFrame]) -> int:
    row = start_row
    for j, val in enumerate(header_cells): worksheet.write(row, j, _xlsx_value(val))
    max_len = max((len(df) for df in blocks.values()), default=0)
    row_blocks_header = row + 1
    start_col = 0
    for crit, _ in CRITERIOS:
        bdf = blocks[crit].copy()
        if len(bdf) < max_len:
            pad = pd.DataFrame({crit: [np.nan]*(max_len - len(bdf)),
                                "numero base": [np.nan]*(max_len - len(bdf)),
                                "Número": [np.nan]*(max_len - len(bdf))})
            bdf = pd.concat([bdf, pad], ignore_index=True)
        worksheet.write(row_blocks_header, start_col + 0, _xlsx_value(crit))
        worksheet.write(row_blocks_header, start_col + 1, _xlsx_value("numero base"))
        worksheet.write(row_blocks_header, start_col + 2, _xlsx_value("Número"))
        col_vals = bdf[crit].tolist(); col_nb = bdf["numero base"].tolist(); col_num = bdf["Número"].tolist()
        for r_off in range(max_len):
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 0, _xlsx_value(col_vals[r_off] if r_off < len(col_vals) else None))
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 1, _xlsx_value(col_nb[r_off] if r_off < len(col_nb) else None))
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 2, _xlsx_value(col_num[r_off] if r_off < len(col_num) else None))
        start_col += 3
    return 1 + 1 + max_len

def _get_header_values(df: pd.DataFrame, sheet_name: str, file_label: str) -> List[str]:
    tam_bloco  = _first_non_null(df, ["Tamanho do Bloco","Tamanho do bloco","Tamanho do bloco: 69"])
    tam_ult    = _first_non_null(df, ["Tamanho do Último Bloco","Tamanho do ultimo bloco"," Tamanho do ultimo bloco: 66"])
    porc_falt  = _first_non_null(df, ["Porcentagem Faltante (%)","Porcentagem Faltante"," Porcentagem Faltante: 4,35%"])
    porc_fmt   = _fmt_percent(porc_falt)
    num_falt   = _first_non_null(df, ["Número de linhas faltantes"," Número de linhas faltantes: 3"])
    return [f"{file_label} :: {sheet_name}",
            f"Tamanho do bloco: {tam_bloco}",
            f"Tamanho do ultimo bloco: {tam_ult}",
            f"Porcentagem Faltante: {porc_fmt}",
            f"Número de linhas faltantes: {num_falt}"]

# ---------- debug ----------
def _debug_dump(df: pd.DataFrame, file_label: str, sheet: str, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    resolver = _build_col_resolver(df)
    cols = {}
    for target in ["máximo grande","mínimo pequeno"]:
        real = _resolve_col(df, target, resolver, SYNONYMS.get(target))
        if real and real in df.columns:
            ser = df[real]
            ser_num = pd.to_numeric(ser.astype(str).str.replace(",",".", regex=False), errors="coerce")
            cols[target] = {
                "encontrada_como": real,
                "n_total": int(len(ser)),
                "n_vazios": int(ser.isna().sum()),
                "n_zero_num": int((ser_num==0).sum()),
                "n_nao_zero_num": int((ser_num!=0).sum()),
                "exemplos": list(ser.head(8).astype(str))
            }
        else:
            cols[target] = {"encontrada_como": None}
    summary = pd.DataFrame(cols).T
    summary.to_csv(outdir / f"DEBUG_{file_label}__{sheet}.csv", index=True, encoding="utf-8-sig")

# ---------- pipeline ----------
def process_files(input_paths: List[str], output_path: str, sheet_name: str = "Planilha Única",
                  zero_as_blank: bool = False, recalc_excel: bool = False, debug: bool = False) -> None:
    inputs: List[str] = []
    for item in input_paths: inputs.extend([p for p in str(item).split(",") if p])

    if recalc_excel:
        try:
            import xlwings as xw
            app = xw.App(visible=False, add_book=False)
            for pth in inputs:
                try:
                    wb = xw.Book(pth); wb.app.calculate(); wb.save(); wb.close()
                except Exception as e:
                    print(f"[aviso] Não consegui recalcular via Excel: {pth} -> {e}")
            app.quit()
        except Exception as e:
            print(f"[aviso] --recalc-excel solicitado, mas não consegui usar xlwings/Excel: {e}")

    writer = pd.ExcelWriter(output_path, engine="xlsxwriter")
    pd.DataFrame([]).to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]

    debug_dir = Path(output_path).with_suffix("").parent / "debug_formula_check"

    current_row = 0
    for wpath in inputs:
        try:
            xls = pd.ExcelFile(wpath); sheets = xls.sheet_names
        except Exception:
            print(f"[aviso] Não consegui abrir '{wpath}'. Pulando."); continue
        for sname in sheets:
            try:
                df = read_sheet_values(wpath, sname)
            except Exception as e:
                print(f"[aviso] Falha ao ler '{wpath}' :: '{sname}': {e}"); continue

            if debug:
                _debug_dump(df, Path(wpath).name, sname, debug_dir)

            header = _get_header_values(df, sname, Path(wpath).name)
            blocks = _build_blocks_side_by_side(df, zero_as_blank=zero_as_blank)
            consumed = _write_section(worksheet, current_row, header, blocks)
            current_row += consumed + 2

    writer.close()

def main():
    ap = argparse.ArgumentParser(description="Planilha única com blocos lado a lado (3 colunas), suportando fórmulas.")
    ap.add_argument("--input","-i", required=True, nargs='+', help="Arquivos de entrada (.xlsx/.xlsm/.ods). Pode repetir -i ou usar vírgula.")
    ap.add_argument("--output","-o", required=True, help="Arquivo XLSX de saída")
    ap.add_argument("--sheet","-s", default="Planilha Única", help="Nome da aba de saída")
    ap.add_argument("--zero-como-vazio", action="store_true", help="Mostra vazio em vez de 0 nas colunas 'máximo grande' e 'mínimo pequeno'.")
    ap.add_argument("--recalc-excel", action="store_true", help="Abre no Excel (xlwings), recalcula e salva antes de ler.")
    ap.add_argument("--debug", action="store_true", help="Gera CSVs por aba mostrando o que foi lido nas colunas-alvo.")
    args = ap.parse_args()
    process_files(args.input, args.output, args.sheet, args.zero_como_vazio, args.recalc_excel, args.debug)
    print(f"OK: gerado '{args.output}'")
if __name__ == "__main__":
    main()
