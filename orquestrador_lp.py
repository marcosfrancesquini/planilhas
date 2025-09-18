#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Orquestrador interativo: cortar_txt -> gerar_planilha -> ordenar
----------------------------------------------------------------
Agora com limpeza de tela (clear/cls) a cada escolha do menu.
"""
from __future__ import annotations

import os, sys, re, unicodedata
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Any

# ======================== UTIL GERAL ============================

def _clear():
    try:
        os.system('cls' if os.name == 'nt' else 'clear')
    except Exception:
        # fallback ANSI
        print("\033c", end="")

def _pause():
    try:
        input("\nPressione ENTER para continuar...")
    except EOFError:
        pass

def _input_str(msg: str, default: Optional[str] = None) -> str:
    s = input(msg).strip()
    if not s and default is not None:
        return default
    return s

def _input_int(msg: str, default: Optional[int] = None) -> int:
    while True:
        s = input(msg).strip()
        if not s and default is not None:
            return default
        try:
            return int(s)
        except Exception:
            print("Valor inválido. Tente novamente.")

def _parse_percent_range(s: str, default: Tuple[float, float] = (3.0, 5.0)) -> Tuple[float, float]:
    s = s.strip()
    if not s:
        return default
    try:
        parts = s.replace("%","").replace(",",".").split("-")
        x = float(parts[0]); y = float(parts[1])
        if x < 0 or y < 0 or x > 100 or y > 100 or x > y:
            raise ValueError
        return (x, y)
    except Exception:
        print("Entrada inválida; usando padrão 3-5%.")
        return default

def _contar_linhas(caminho: str, encoding: str = "utf-8") -> int:
    total = 0
    with open(caminho, "r", encoding=encoding, newline="") as f:
        for _ in f: total += 1
    return total

def _tokenizar_numeros(linha: str) -> List[int]:
    # aceita separadores espaço, vírgula, ponto-e-vírgula, tab etc
    linha = linha.strip()
    if not linha:
        return []
    linha = linha.replace(",", " ").replace(";", " ").replace("\t", " ")
    nums: List[int] = []
    for p in linha.split():
        try:
            n = int(p)
        except Exception:
            continue
        if 0 <= n <= 99:
            nums.append(n)
    return nums

def _formatar_com_virgulas(nums: List[int]) -> str:
    # mantém sem zero à esquerda, separados por vírgula e espaço
    return ", ".join(str(n) for n in nums)

def _ler_unica_linha_numeros(caminho: str) -> List[int]:
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            for raw in f:
                return _tokenizar_numeros(raw)
        return []
    except Exception:
        return []

# ======================== 1) CORTAR TXT =========================

def cortar_txt(caminho: str, corte: int, encoding: str = "utf-8") -> Tuple[str, str, int]:
    """
    Retorna (arquivo_primeiras, arquivo_linha_unica, total_linhas).
    - Salva os arquivos já com números separados por vírgula.
    """
    total = _contar_linhas(caminho, encoding=encoding)
    print(f"Total de linhas: {total}")
    if corte < 1 or corte >= total:
        raise ValueError("valor de corte inválido (precisa ser 1..total-1).")

    base_dir = Path(caminho).resolve().parent
    stem = Path(caminho).resolve().stem
    arq_primeiras = base_dir / f"{stem}_primeiras_{corte}.txt"
    arq_linha_unica = base_dir / f"{stem}_linha_{corte+1}.txt"

    linha_n_mais_1_nums: Optional[List[int]] = None
    with open(caminho, "r", encoding=encoding, newline="") as fin, \
         open(arq_primeiras, "w", encoding=encoding, newline="") as fout_primeiras:
        for idx, linha in enumerate(fin, start=1):
            if idx <= corte:
                nums = _tokenizar_numeros(linha)
                fout_primeiras.write(_formatar_com_virgulas(nums) + "\n")
            elif idx == corte + 1:
                linha_n_mais_1_nums = _tokenizar_numeros(linha)
            else:
                pass

    with open(arq_linha_unica, "w", encoding=encoding, newline="") as f:
        if linha_n_mais_1_nums is not None:
            f.write(_formatar_com_virgulas(linha_n_mais_1_nums) + "\n")

    print("✅ Corte concluído:")
    print(f" - Linha {corte+1} (uma linha) -> {arq_linha_unica}")
    print(f" - Primeiras {corte} linhas     -> {arq_primeiras}")
    return str(arq_primeiras), str(arq_linha_unica), total

# ======================== 2) GERAR PLANILHA =====================

from typing import Iterable
import pandas as pd

def ler_sorteios_txt(caminho: str) -> List[List[int]]:
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")
    sorteios: List[List[int]] = []
    with open(caminho, 'r', encoding='utf-8') as f:
        for idx, raw in enumerate(f, 1):
            raw = raw.strip()
            if not raw:
                continue
            # aceita vírgula, ponto-e-vírgula, tab e espaço
            raw = raw.replace('\t', ' ').replace(';', ' ').replace(',', ' ')
            partes = [p for p in raw.split(' ') if p != '']
            linha: List[int] = []
            for p in partes:
                try:
                    n = int(p)
                except ValueError:
                    print(f"[AVISO] Linha {idx}: token inválido '{p}' – ignorado.", file=sys.stderr)
                    continue
                if 0 <= n <= 99:
                    linha.append(n)
            if linha:
                sorteios.append(linha)
    if not sorteios:
        raise ValueError("Nenhuma linha válida encontrada no arquivo.")
    return sorteios

def analise_por_bloco(linhas: List[List[int]], K: int):
    N = len(linhas)
    C = N // K
    r = N - C*K
    dados: Dict[int, tuple[int, int, int]] = {}
    if C == 0:
        for n in range(100):
            freq_incomp = 0
            for i in range(0, r):
                freq_incomp += sum(1 for x in linhas[i] if x == n)
            dados[n] = (0, 0, freq_incomp)
        return C, r, dados

    blocos_freq: List[List[int]] = []
    for j in range(C):
        i0 = j*K
        i1 = (j+1)*K
        hist = [0]*100
        for i in range(i0, i1):
            for x in linhas[i]:
                if 0 <= x <= 99:
                    hist[x] += 1
        blocos_freq.append(hist)

    hist_incomp = [0]*100
    if r > 0:
        i0 = C*K
        i1 = N
        for i in range(i0, i1):
            for x in linhas[i]:
                if 0 <= x <= 99:
                    hist_incomp[x] += 1

    for n in range(100):
        col = [blocos_freq[j][n] for j in range(C)]
        max_c = max(col) if col else 0
        min_c = min(col) if col else 0
        dados[n] = (max_c, min_c, hist_incomp[n] if r > 0 else 0)

    return C, r, dados

def gerar_planilhas(
    caminho_txt: str,
    faixa_percent: Tuple[float, float] = (3.0, 5.0),
    max_x: int = 10,
    min_x: int = 0,
    numeros_base: List[int] | None = None,
    saida_xlsx: str | None = None
) -> str:
    if numeros_base is None:
        numeros_base = []
    linhas = ler_sorteios_txt(caminho_txt)
    N = len(linhas)

    x, y = faixa_percent
    if x < 0 or y < 0 or x > 100 or y > 100 or x > y:
        raise ValueError("Faixa de porcentagem inválida. Use, por exemplo, 3 a 5.")

    if saida_xlsx is None:
        base = os.path.splitext(os.path.basename(caminho_txt))[0]
        saida_xlsx = f"analise_lotomania_{base}.xlsx"

    with pd.ExcelWriter(saida_xlsx, engine="xlsxwriter") as writer:
        for K in range(2, N+1):
            C = N // K
            r = N - C*K
            if r == 0:
                continue
            faltantes = K - r
            pct_faltante = (faltantes / K) * 100.0
            if pct_faltante + 1e-9 < x or pct_faltante - 1e-9 > y:
                continue

            C, r, dados = analise_por_bloco(linhas, K)

            registros = []
            for num in range(100):
                max_c, min_c, freq_r = dados[num]
                max_vezes = max_c - freq_r
                min_vezes = freq_r - min_c
                registro = {
                    "Tamanho do Bloco": K,
                    "Tamanho do Último Bloco": r,
                    "Porcentagem Faltante (%)": round(pct_faltante, 2),
                    "Número de linhas faltantes": faltantes,
                    "Número": f"{num:02d}",
                    "Máxima Freq. nos Blocos Completos": max_c,
                    "Mínima Freq. nos Blocos Completos": min_c,
                    "Frequência no Último Bloco Incompleto": freq_r,
                    "Maximo de vezes": max_vezes,
                    "Minimo de vezes": min_vezes,
                    "máximo grande": "",
                    "mínimo pequeno": "",
                    "numero base": (num if num in numeros_base else "")
                }
                registros.append(registro)

            df = pd.DataFrame(registros, columns=[
                "Tamanho do Bloco","Tamanho do Último Bloco","Porcentagem Faltante (%)",
                "Número de linhas faltantes","Número",
                "Máxima Freq. nos Blocos Completos","Mínima Freq. nos Blocos Completos",
                "Frequência no Último Bloco Incompleto",
                "Maximo de vezes","Minimo de vezes",
                "máximo grande","mínimo pequeno","numero base"
            ])

            sheet_name = f"K={K} ({pct_faltante:.2f}%)"
            if len(sheet_name) > 31:
                sheet_name = f"K={K}"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            first_row = 2
            last_row = len(df) + 1
            for row in range(first_row, last_row + 1):
                formula_k = f'=IF(I{row}>={max_x},E{row},"")'
                formula_l = f'=IF(J{row}<={min_x},E{row},"")'
                ws.write_formula(row - 1, 10, formula_k)  # K
                ws.write_formula(row - 1, 11, formula_l)  # L

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.set_column(0, 1, 22)
            ws.set_column(2, 3, 26)
            ws.set_column(4, 4, 10)
            ws.set_column(5, 8, 30)
            ws.set_column(9, 11, 18)
            ws.set_column(12, 12, 14)

    return os.path.abspath(saida_xlsx)

def parse_lista_numeros(txt: str) -> List[int]:
    if not txt.strip():
        return []
    itens = []
    for p in txt.replace(',', ' ').split():
        try:
            n = int(p)
        except ValueError:
            continue
        if 0 <= n <= 99:
            itens.append(n)
    return sorted(set(itens))

# ======================== 3) ORDENAR ============================

import argparse, numpy as np

CRITERIOS: List[Tuple[str, bool]] = [
    ("Máxima Freq. nos Blocos Completos", False),
    ("Mínima Freq. nos Blocos Completos", True),
    ("Frequência no Último Bloco Incompleto", False),
    ("Maximo de vezes", False),
    ("Minimo de vezes", True),
    ("máximo grande", False),
    ("mínimo pequeno", True),
]

SYNONYMS: Dict[str, List[str]] = {
    "Máxima Freq. nos Blocos Completos": ["Maxima Freq. nos Blocos Completos","Maxima Freq nos Blocos Completos","Máxima Freq nos Blocos Completos"],
    "Mínima Freq. nos Blocos Completos": ["Minima Freq. nos Blocos Completos","Minima Freq nos Blocos Completos","Mínima Freq nos Blocos Completos"],
    "Frequência no Último Bloco Incompleto": ["Frequencia no Ultimo Bloco Incompleto","Frequencia no Último Bloco Incompleto"],
    "Maximo de vezes": ["Máximo de vezes","Maximo de Vezes"],
    "Minimo de vezes": ["Mínimo de vezes","Minimo de Vezes"],
    "máximo grande": ["Maximo grande","maximo grande"],
    "mínimo pequeno": ["Minimo pequeno","minimo pequeno"],
}

def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def _build_col_resolver(df: pd.DataFrame) -> Dict[str, str]:
    return {_norm_text(c): c for c in df.columns}

def _resolve_col(df: pd.DataFrame, canonical: str, resolver: Dict[str, str], extras: Optional[List[str]] = None) -> Optional[str]:
    candidates = [canonical] + (extras or [])
    for cand in candidates:
        key = _norm_text(cand)
        if key in resolver:
            return resolver[key]
    key = _norm_text(canonical).replace("freq","frequencia")
    return resolver.get(key)

def _first_non_null(df: pd.DataFrame, candidates: List[str]) -> Any:
    for c in candidates:
        if c in df.columns:
            s = pd.Series(df[c]).dropna()
            if len(s) > 0:
                return s.iloc[0]
    return np.nan

def _fmt_percent(val: Any) -> str:
    if pd.isna(val): return ""
    try:
        s = str(val).strip().replace("%","").replace(",",".")
        return f"{float(s):.2f}%"
    except Exception:
        return str(val)

def _pick_numero_base(df: pd.DataFrame) -> pd.Series:
    if "numero base" in df.columns: return df["numero base"]
    for alt in ["Número","numero","numero_base","número base"]:
        if alt in df.columns:
            s = df[alt].copy(); s.name = "numero base"; return s
    return pd.Series(df.index + 1, index=df.index, name="numero base")

def _pick_numero_col(df: pd.DataFrame) -> pd.Series:
    for alt in ["Número","numero","num","Numero","NÚMERO","número"]:
        if alt in df.columns:
            s = df[alt].copy(); s.name = "Número"; return s
    nb = _pick_numero_base(df); s = nb.copy(); s.name = "Número"; return s

def _coerce_numeric_col(series: pd.Series) -> pd.Series:
    if series.dtype != object: return series
    try:
        return pd.to_numeric(series.astype(str).str.replace("%","",regex=False).str.replace(",",".",regex=False).str.strip(), errors="coerce")
    except Exception:
        return series

def read_sheet_values(path: str, sheet_name: str) -> pd.DataFrame:
    p = Path(path); ext = p.suffix.lower()
    if ext in [".xlsx",".xlsm"]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames: return pd.DataFrame()
            ws = wb[sheet_name]
            rows = list(ws.values)
            if not rows: return pd.DataFrame()
            header = [f"Unnamed: {i}" if (h in (None,"")) else h for i,h in enumerate(rows[0])]
            data = rows[1:]
            return pd.DataFrame(data, columns=header)
        except Exception:
            return pd.read_excel(path, sheet_name=sheet_name)
    else:
        return pd.read_excel(path, sheet_name=sheet_name)

def _build_blocks_side_by_side(df: pd.DataFrame, zero_as_blank: bool) -> Dict[str, pd.DataFrame]:
    nb = _pick_numero_base(df)
    numcol = _pick_numero_col(df)
    blocks: Dict[str, pd.DataFrame] = {}
    resolver = _build_col_resolver(df)

    for crit, asc in CRITERIOS:
        real_col = _resolve_col(df, crit, resolver, SYNONYMS.get(crit))
        if real_col is None:
            ord_df = pd.DataFrame({crit: pd.Series(dtype=object), "numero base": pd.Series(dtype=object), "Número": pd.Series(dtype=object)})
        else:
            orig = df[real_col].copy()
            if zero_as_blank and crit in ["máximo grande","mínimo pequeno"]:
                if pd.api.types.is_numeric_dtype(orig):
                    import numpy as _np
                    orig = orig.mask(orig == 0, other=_np.nan)
                else:
                    orig = orig.mask(orig.astype(str).str.strip().isin({"0","0.0","0,0"}), other=float("nan"))
            key = _coerce_numeric_col(df[real_col])
            ord_df = pd.DataFrame({crit: orig, "numero base": nb, "Número": numcol, "__key__": key})
            ord_df = ord_df.sort_values(by="__key__", ascending=asc, kind="mergesort", na_position="last").drop(columns=["__key__"])
        blocks[crit] = ord_df.reset_index(drop=True)
    return blocks

def _xlsx_value(v):
    try:
        import pandas as _pd, numpy as _np
        if v is None: return None
        if isinstance(v,(float,int)) and (_pd.isna(v) or (isinstance(v,float) and (_np.isnan(v) or _np.isinf(v)))): return None
        if _pd.isna(v): return None
        return v
    except Exception:
        return v

def _write_section(worksheet, start_row: int, header_cells: List[str], blocks: Dict[str, pd.DataFrame]) -> int:
    row = start_row
    for j, val in enumerate(header_cells): worksheet.write(row, j, _xlsx_value(val))
    max_len = max((len(df) for df in blocks.values()), default=0)
    row_blocks_header = row + 1
    start_col = 0
    for crit, _ in CRITERIOS:
        bdf = blocks[crit].copy()
        if len(bdf) < max_len:
            import pandas as _pd, numpy as _np
            pad = _pd.DataFrame({crit: [_np.nan]*(max_len - len(bdf)),
                                "numero base": [_np.nan]*(max_len - len(bdf)),
                                "Número": [_np.nan]*(max_len - len(bdf))})
            bdf = pd.concat([bdf, pad], ignore_index=True)
        worksheet.write(row_blocks_header, start_col + 0, _xlsx_value(crit))
        worksheet.write(row_blocks_header, start_col + 1, _xlsx_value("numero base"))
        worksheet.write(row_blocks_header, start_col + 2, _xlsx_value("Número"))
        col_vals = bdf[crit].tolist(); col_nb = bdf["numero base"].tolist(); col_num = bdf["Número"].tolist()
        for r_off in range(max_len):
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 0, _xlsx_value(col_vals[r_off] if r_off < len(col_vals) else None))
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 1, _xlsx_value(col_nb[r_off] if r_off < len(col_nb) else None))
            worksheet.write(row_blocks_header + 1 + r_off, start_col + 2, _xlsx_value(col_num[r_off] if r_off < len(col_num) else None))
        start_col += 3
    return 1 + 1 + max_len

def _get_header_values(df: pd.DataFrame, sheet_name: str, file_label: str) -> List[str]:
    tam_bloco  = _first_non_null(df, ["Tamanho do Bloco","Tamanho do bloco","Tamanho do bloco: 69"])
    tam_ult    = _first_non_null(df, ["Tamanho do Último Bloco","Tamanho do ultimo bloco"," Tamanho do ultimo bloco: 66"])
    porc_falt  = _first_non_null(df, ["Porcentagem Faltante (%)","Porcentagem Faltante"," Porcentagem Faltante: 4,35%"])
    porc_fmt   = _fmt_percent(porc_falt)
    num_falt   = _first_non_null(df, ["Número de linhas faltantes"," Número de linhas faltantes: 3"])
    return [f"{file_label} :: {sheet_name}",
            f"Tamanho do bloco: {tam_bloco}",
            f"Tamanho do ultimo bloco: {tam_ult}",
            f"Porcentagem Faltante: {porc_fmt}",
            f"Número de linhas faltantes: {num_falt}"]

def process_files_ordenar(input_paths: List[str], output_path: str, sheet_name: str = "Planilha Única",
                  zero_as_blank: bool = False, recalc_excel: bool = False, debug: bool = False) -> None:
    inputs: List[str] = []
    for item in input_paths: inputs.extend([p for p in str(item).split(",") if p])

    if recalc_excel:
        try:
            import xlwings as xw
            app = xw.App(visible=False, add_book=False)
            for pth in inputs:
                try:
                    wb = xw.Book(pth); wb.app.calculate(); wb.save(); wb.close()
                except Exception as e:
                    print(f"[aviso] Não consegui recalcular via Excel: {pth} -> {e}")
            app.quit()
        except Exception as e:
            print(f"[aviso] --recalc-excel solicitado, mas não consegui usar xlwings/Excel: {e}")

    writer = pd.ExcelWriter(output_path, engine="xlsxwriter")
    pd.DataFrame([]).to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]

    current_row = 0
    for wpath in inputs:
        try:
            xls = pd.ExcelFile(wpath); sheets = xls.sheet_names
        except Exception:
            print(f"[aviso] Não consegui abrir '{wpath}'. Pulando."); continue
        for sname in sheets:
            try:
                df = read_sheet_values(wpath, sname)
            except Exception as e:
                print(f"[aviso] Falha ao ler '{wpath}' :: '{sname}': {e}"); continue

            header = _get_header_values(df, sname, Path(wpath).name)
            blocks = _build_blocks_side_by_side(df, zero_as_blank=zero_as_blank)
            consumed = _write_section(worksheet, current_row, header, blocks)
            current_row += consumed + 2

    writer.close()

# ======================== MENU / PIPELINE =======================

def _detectar_n_do_txt(caminho_txt: str) -> int:
    # retorna número de linhas do arquivo TXT
    return _contar_linhas(caminho_txt, encoding="utf-8")

def pipeline_completa():
    _clear()
    print("\n=== Pipeline Completa: cortar -> planilha -> ordenar ===")
    caminho = _input_str("Caminho do arquivo .txt original: ")
    corte = _input_int("Número da linha onde cortar (N): ")
    arq_primeiras, arq_unica, total = cortar_txt(caminho, corte)

    # Opcional: usar a linha única como números base
    base_nums = _ler_unica_linha_numeros(arq_unica)
    usar_base = _input_str(f"Usar a linha {corte+1} como Números base? (s/n) [n]: ", default="n").lower().startswith("s")
    numeros_base = base_nums if usar_base else []

    # Parâmetros da planilha
    faixa = _parse_percent_range(_input_str("Faixa % faltante para último bloco (x-y) [3-5]: "))
    max_x = _input_int("Limiar 'Máximo Maior' (padrão 10): ", default=10)
    min_x = _input_int("Limiar 'Mínimo Menor' (padrão 0): ", default=0)

    saida_planilha = str(Path(arq_primeiras).with_name(f"Planilha_{corte}.xlsx"))
    caminho_saida = gerar_planilhas(
        caminho_txt=arq_primeiras,
        faixa_percent=faixa,
        max_x=max_x,
        min_x=min_x,
        numeros_base=numeros_base,
        saida_xlsx=saida_planilha
    )
    print(f"✅ Planilha gerada: {caminho_saida}")

    # Ordenar
    out_ordenado = str(Path(caminho_saida).with_name(f"ordenado{corte}.xlsx"))
    process_files_ordenar([caminho_saida], out_ordenado, sheet_name="Planilha Única", zero_as_blank=False, recalc_excel=False, debug=False)
    print(f"✅ Arquivo ordenado: {out_ordenado}")
    _pause()

def so_cortar():
    _clear()
    print("\n=== Só Cortar TXT ===")
    caminho = _input_str("Caminho do arquivo .txt original: ")
    corte = _input_int("Número da linha onde cortar (N): ")
    cortar_txt(caminho, corte)
    _pause()

def so_gerar_planilha():
    _clear()
    print("\n=== Só Gerar Planilha (de um TXT) ===")
    caminho_txt = _input_str("Arquivo TXT com sorteios: ")
    faixa = _parse_percent_range(_input_str("Faixa % faltante para último bloco (x-y) [3-5]: "))
    max_x = _input_int("Limiar 'Máximo Maior' (padrão 10): ", default=10)
    min_x = _input_int("Limiar 'Mínimo Menor' (padrão 0): ", default=0)

    usar_arq_unico = _input_str("Usar um arquivo de 'uma linha' como Números base? (s/n) [n]: ", default="n").lower().startswith("s")
    numeros_base: List[int] = []
    if usar_arq_unico:
        arq_unico = _input_str("Caminho do arquivo de uma linha: ")
        numeros_base = _ler_unica_linha_numeros(arq_unico)
        print(f"Números base detectados: {numeros_base}")

    n = _detectar_n_do_txt(caminho_txt)
    saida_planilha = str(Path(caminho_txt).with_name(f"Planilha_{n}.xlsx"))
    caminho_saida = gerar_planilhas(
        caminho_txt=caminho_txt,
        faixa_percent=faixa,
        max_x=max_x,
        min_x=min_x,
        numeros_base=numeros_base,
        saida_xlsx=saida_planilha
    )
    print(f"✅ Planilha gerada: {caminho_saida}")
    _pause()

def so_ordenar():
    _clear()
    print("\n=== Só Ordenar (de um XLSX) ===")
    input_xlsx = _input_str("Arquivo Excel de entrada (.xlsx): ")
    # tenta extrair N do nome Planilha_N.xlsx; se não der, usa 'ordenado.xlsx'
    m = re.search(r"Planilha[_\-](\d+)\.xlsx$", os.path.basename(input_xlsx), flags=re.IGNORECASE)
    if m:
        n = m.group(1)
        out = str(Path(input_xlsx).with_name(f"ordenado{n}.xlsx"))
    else:
        out = str(Path(input_xlsx).with_name("ordenado.xlsx"))
    process_files_ordenar([input_xlsx], out, sheet_name="Planilha Única", zero_as_blank=False, recalc_excel=False, debug=False)
    print(f"✅ Arquivo ordenado: {out}")
    _pause()

def menu():
    while True:
        _clear()
        print("================ MENU ================")
        print("1) Pipeline completa (cortar -> planilha -> ordenar)")
        print("2) Só cortar TXT")
        print("3) Só gerar planilha (a partir de TXT)")
        print("4) Só ordenar (a partir de XLSX)")
        print("0) Sair")
        op = _input_str("Escolha: ")
        if op == "1":
            try:
                pipeline_completa()
            except Exception as e:
                print(f"\nERRO na pipeline: {e}")
                _pause()
        elif op == "2":
            try:
                so_cortar()
            except Exception as e:
                print(f"\nERRO no corte: {e}")
                _pause()
        elif op == "3":
            try:
                so_gerar_planilha()
            except Exception as e:
                print(f"\nERRO ao gerar planilha: {e}")
                _pause()
        elif op == "4":
            try:
                so_ordenar()
            except Exception as e:
                print(f"\nERRO ao ordenar: {e}")
                _pause()
        elif op == "0":
            print("Até mais!")
            break
        else:
            print("Opção inválida.")
            _pause()

if __name__ == "__main__":
    menu()
